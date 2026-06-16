import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import pandas as pd
import os
import shutil

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'projet_s2'

clients = [['Rakoto', 'Jean', 0],
         ['Rabe', 'Marie', -15000],
         ['Andry', 'Paul', -8000],
         ['Rasoa', 'Clara', 0],
         ['Raivo', 'Marc', -32000],
         ['Randria', 'Sophie', 0],
         ['Ratsima', 'Luc', -5000]]

entetes = ['nom', 'prenom', 'solde']

#pour l'entete:
for col, entete in enumerate(entetes, start=1):
    entete_cell = ws.cell(row=1, column=col, value=entete)
    entete_cell.font = Font(bold=True, color="FFFFFF")
    entete_cell.fill = PatternFill("solid", fgColor='1E3A5F')

#entrer des lignes :
for row, ligne in enumerate(clients, start=2):
    for col, valeur in enumerate(ligne, start=1):
        ws.cell(row= row, column = col, value = valeur)

wb.save('projet_s2.xlsx')

df = pd.read_excel('projet_s2.xlsx')

#filtre des clients dont le solde < 0:
filtre_client = df[df['solde'] < 0]
print(filtre_client)

#sauvegader les resultats en excel : 
filtre_client.to_excel('clients_a_relancer.xlsx', index=False)
print('sauvegarde reussie: clients_a_relancer.xlsx')



#rangement du dossier :
type_fichier = {
    '.xlsx' : 'EXCEL',
    '.csv' : 'CSV',
    '.txt' : 'TEXTE',
    '.py' : 'SCRIPT PYTHON',
    '.docx' : 'WORD'
    }

print("--DEBUT DE L'AUTOMATISATION--")
dossier = os.getcwd()
contenus = os.listdir(dossier)
# chemin_contenus = os.path.join(dossier, contenus)

for element in contenus:
    #chemin de chaque element
    chemin_element = os.path.join(dossier, element)
    if os.path.isdir(element):
        continue
    else:
        nom, extension = os.path.splitext(element)
        if extension in type_fichier:
            chemin_dossier = os.path.join(dossier, type_fichier[extension])
            os.makedirs(chemin_dossier, exist_ok=True)
            shutil.move(chemin_element, chemin_dossier)
            print("\\ AUTOMATISATION DE FICHIER REUSSIT AVEC SUCCES //")

        else : 
            print('fichier non reconnue, trier dans dossier inconnue')
            if not os.path.exists(dossier_inconnue):
                dossier_inconnue = os.path.join(dossier, 'dossier inconnue')
                os.makedirs(dossier_inconnue, exist_ok=True)
                shutil.move(chemin_element, dossier_inconnue)


            











